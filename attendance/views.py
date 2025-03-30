from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
import jdatetime
from django.contrib.auth.decorators import login_required
from .models import AttendaceRecord, AttendanceRequest
from datetime import timedelta
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
#get user model
from django.contrib.auth import get_user_model
User = get_user_model()
from django.http import FileResponse
import tempfile
from lunch.models import Lunch



@login_required
def attendance(request):
    user = request.user
    today = jdatetime.date.today()

    # Try to get the attendance record for today
    try:
        record = AttendaceRecord.objects.get(user=user, date=today)
    except AttendaceRecord.DoesNotExist:
        record = None

    if request.method == 'POST':
        if 'checkin' in request.POST:
            checkin_time_str = request.POST.get('checkin')
            check_in = jdatetime.datetime.strptime(checkin_time_str, '%H:%M').time()

            if record:
                messages.error(request, "شما ساعت ورود خود را ثبت کرده اید.")
            else:
                # Create a new record if it doesn't exist (first check-in)
                AttendaceRecord.objects.create(user=user, check_in=check_in, date=today)
                messages.success(request, "ساعت ورود با موفقیت ثبت شد.")
                return redirect('attendance')  # Redirect to the same view after processing

        if 'checkout' in request.POST:
            checkout_time_str = request.POST.get('checkout')
            check_out = jdatetime.datetime.strptime(checkout_time_str, '%H:%M').time()

            if record and record.check_out:
                messages.error(request, "شما ساعت خروج خود را ثبت کرده اید.")
            elif record and record.check_in:
                # Set the checkout time if the check-in time is already set
                record.check_out = check_out
                record.save()
                messages.success(request, "ساعت خروج با موفقیت ثبت شد.")
                return redirect('attendance')  # Redirect to the same view after processing

    # Pass the existing check-in/check-out times and other context to the template
    context = {
        'check_in': record.check_in if record else None,
        'check_out': record.check_out if record else None,
        'day_name': today.strftime('%A'),
        'today': today
    }

    return render(request, 'attendance/attendance.html', context)

from django.views.decorators.csrf import csrf_exempt
import json
from django.http import JsonResponse
from datetime import datetime


@csrf_exempt
# api for setting checkin and checkout from rfidreader
def attendance_api(request):
    if request.method == "POST":
        data = json.loads(request.body)
        rfid = data.get("rfid")
        received_time = data.get("time")
        gtime = datetime.fromisoformat(received_time).replace(tzinfo=None)
        time = jdatetime.datetime.fromgregorian(datetime=gtime).time()
        print(time)
        try:
            user = User.objects.get(rfid=rfid)
        except User.DoesNotExist:
            return JsonResponse({"status": "error", "message": "user not found"})
        today = jdatetime.date.today()
        try:
            record = AttendaceRecord.objects.get(user=user, date=today)
        except AttendaceRecord.DoesNotExist:
            record = None

        if record:
            if record.check_in and record.check_out:
                return JsonResponse(
                    {
                        "status": "error",
                        "message": "you have already checked in and out today",
                    }
                )
            if time and record.check_in and not record.check_out:
                check_out = time
                record.check_out = check_out
                record.save()
                return JsonResponse(
                    {
                        "status": "success",
                        "message": "checkout time set successfully",
                    }
                )
            else:
                return JsonResponse({"status": "error", "message": "invalid request"})
        else:
            if time:
                check_in = time
                AttendaceRecord.objects.create(user=user, check_in=check_in, date=today)
                return JsonResponse(
                    {"status": "success", "message": "checkin time set successfully"}
                )
            else:
                return JsonResponse({"status": "error", "message": "invalid request"})




def attendance_list(request):
    # Get the attendance records for the logged-in user
    attendances = AttendaceRecord.objects.filter(user=request.user).order_by('-date')
    return render(request, 'attendance/attendance_list.html', {'attendances': attendances})


@login_required
def change_attendance(request, attendance_id=None):
    # Retrieve the attendance record if an ID is provided
    if attendance_id:
        attendance = get_object_or_404(AttendaceRecord, id=attendance_id, user=request.user)
    else:
        attendance = None  # For new attendance requests

    if request.method == 'POST':
        # Manually retrieve data from the POST request
        requested_check_in = request.POST.get('requested_check_in')
        requested_check_out = request.POST.get('requested_check_out')
        date_str = request.POST.get('date')  # Expecting a date string in the format you define
        print(date_str)
        request_reason = request.POST.get('request_reason')

        if attendance:
        # Validate the input data
            if not requested_check_in or not requested_check_out or not request_reason:
                messages.error(request, "لطفا تمام فیلدها را پر کنید.")
                return render(request, 'attendance/attendance_edit.html', {'attendance': attendance})
            requested_date = attendance.date  # Use the existing date for the attendance record
        else:
            if not requested_check_in or not requested_check_out or not date_str or not request_reason:
                messages.error(request, "لطفا تمام فیلدها را پر کنید.")
                return render(request, 'attendance/attendance_edit.html', {'attendance': attendance})

        # Convert the date string to a jdatetime object
            try:
                requested_date = jdatetime.datetime.strptime(date_str, '%Y/%m/%d').date()  # Adjust format as needed
            except ValueError:
                messages.error(request, "فرمت تاریخ نادرست است.")
                return render(request, 'attendance/attendance_edit.html')

        # Create a new AttendanceRequest
        attendance_request = AttendanceRequest(
            user=request.user,
            attendance=attendance,  # This will be None for new requests
            requested_check_in=jdatetime.datetime.strptime(requested_check_in, '%H:%M').time(),
            requested_check_out=jdatetime.datetime.strptime(requested_check_out, '%H:%M').time(),
            date=requested_date,
            request_reason=request_reason
        )
        attendance_request.save()  # Save the change request
        messages.success(request, "درخواست تغییر با موفقیت ثبت شد.")  # Success message
        return redirect('user_requests')  # Redirect to the attendance list

    return render(request, 'attendance/attendance_edit.html', {'attendance': attendance})

def user_requests(request):
    user_requests = AttendanceRequest.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'attendance/attendance_requests.html', {'user_requests': user_requests})

@login_required
def report_attendance(request):
    if request.user.is_superuser:
        # Create a new workbook
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet

        # Get unique users with attendance records
        users = User.objects.all()

        for user in users:
            # Create a new sheet for each user
            sheet = workbook.create_sheet(title=f"{user.last_name}")

            # Header row
            headers = ["روز", "تاریخ", "ساعت ورود", "ساعت خروج", "کارکرد", "درآمد"]
            sheet.append(headers)
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                sheet[f"{column_letter}1"].font = Font(bold=True)
                sheet[f"{column_letter}1"].alignment = Alignment(horizontal="center")

            # Get attendance records for the current month
            start_date, end_date = AttendaceRecord.current_month_date_range()
            records = AttendaceRecord.objects.filter(user=user, date__gte=start_date, date__lte=end_date)

            # Fill rows with attendance data
            total_hours = timedelta()
            total_price = 0
            jdatetime.set_locale(jdatetime.FA_LOCALE)
            for record in records:
                day_name = record.date.strftime("%A")
                total_time = record.duration()
                daily_price = record.daily_total_price()
                total_hours += total_time
                total_price += daily_price

                sheet.append([
                    day_name,
                    record.date.strftime("%Y-%m-%d"),
                    record.check_in.strftime("%H:%M"),
                    record.check_out.strftime("%H:%M") if record.check_out else "N/A",
                    str(total_time),
                    round(daily_price, 2),
                ])
            reserved_lunches = Lunch.objects.filter(user=user, date__gte=start_date, date__lte=end_date).count()
            # Add totals row
            sheet.append([])
            sheet.append(["", "", "", "جمع", total_hours, round(total_price, 2)])
            sheet.append(["", "", "", "تعداد ناهار", reserved_lunches])
            sheet.append(["","","","هزینه ناهار", reserved_lunches * 65000])
            sheet.append(["","","","مجموع", total_price - (reserved_lunches * 65000)])

            for col_num in range(1, 7):
                column_letter = get_column_letter(col_num)
                sheet[f"{column_letter}{sheet.max_row}"].font = Font(bold=True)

        # Save workbook to a temporary file and return response
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            response = FileResponse(open(tmp.name, 'rb'), as_attachment=True, filename="Attendance_Report.xlsx")

        return response
